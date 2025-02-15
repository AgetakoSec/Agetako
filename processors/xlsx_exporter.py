from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from config.site_config import EXCEL_FILE, SITE_CONFIG


def save_filtered_articles_to_xlsx(filtered_articles):
    """
    フィルタリングされた記事をXLSXファイルに保存し、重複のない新しい記事を追加する。
    """
    file_name = Path(EXCEL_FILE)
    table_name = "Vulnerability"

    # 現在の時刻
    current_time = datetime.now()

    # 必須列
    required_columns = [
        "Date",
        "Site",
        "SiteLink",
        "Title",
        "Description",
        "CVE",
        "CVSS",
        "link",
        "Post",
    ]

    # 既存データの読み込みまたは新規作成
    if file_name.exists():
        workbook = load_workbook(file_name)
        if table_name in workbook.sheetnames:
            sheet = workbook[table_name]
            data = list(sheet.values)
            if data:
                columns = [str(col) for col in data[0]]  # 列名を取得
                rows = data[1:]
                existing_data = pd.DataFrame(rows, columns=columns)
            else:
                existing_data = pd.DataFrame(columns=required_columns)
        else:
            sheet = workbook.create_sheet(title=table_name)
            existing_data = pd.DataFrame(columns=required_columns)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = table_name
        existing_data = pd.DataFrame(columns=required_columns)

    # 既存データをリスト形式に変換
    if not existing_data.empty:
        existing_data["Date"] = pd.to_datetime(existing_data["Date"], errors="coerce")
        existing_links = set(existing_data["link"].dropna().tolist())
    else:
        existing_links = set()

    # 新しい記事を追加
    new_articles = [
        article
        for article in filtered_articles
        if article["link"] not in existing_links
    ]

    if not new_articles:
        print("No new articles to process.")
        return

    # DataFrame作成
    new_articles_df = pd.DataFrame(
        [
            {
                "Date": article.get("Date", ""),
                "Site": article.get("SiteName", ""),
                "SiteLink": SITE_CONFIG.get(article.get("SiteName", ""), {}).get(
                    "url", ""
                ),
                "Title": article.get("Title", ""),
                "Description": article.get("Description", ""),
                "CVE": article.get("CVE", ""),
                "CVSS": article.get("CVSS", ""),
                "link": article.get("link", ""),
                "Post": "未投稿",
            }
            for article in filtered_articles
        ]
    )

    # 全データを統合
    all_articles_df = pd.concat([existing_data, new_articles_df], ignore_index=True)

    # 7日以上前の記事を削除
    cutoff_date = current_time - timedelta(days=7)
    all_articles_df["Date"] = pd.to_datetime(all_articles_df["Date"], errors="coerce")
    all_articles_df = all_articles_df[all_articles_df["Date"] >= cutoff_date]

    # 重複行を削除
    all_articles_df = all_articles_df.drop_duplicates(subset=["link"])

    # 降順でソート
    all_articles_df = all_articles_df.sort_values(by=["Date", "Site"], ascending=False)

    # 書き込み
    with pd.ExcelWriter(file_name, engine="openpyxl", mode="w") as writer:
        all_articles_df.to_excel(writer, sheet_name=table_name, index=False)

    # テーブル作成
    workbook = load_workbook(file_name)
    sheet = workbook[table_name]

    # テーブルの範囲設定
    min_row, max_row = 1, sheet.max_row
    min_col, max_col = 1, sheet.max_column
    table = Table(
        displayName=table_name, ref=f"A{min_row}:I{max_row}"
    )  # ArticleNum列を含むように調整
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style
    sheet.add_table(table)

    workbook.save(file_name)
    print(f"Updated {file_name} with {len(new_articles)} new articles.")
